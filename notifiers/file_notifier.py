import datetime

from notifiers.notifier import Notifier
import os.path
import openpyxl

APARTMENTS_WORKSHEET = 'apartments'


class FileNotifier(Notifier):
    def __init__(self, file_path):
        self.file_path = file_path
        if os.path.isfile(file_path):
            self.wb = openpyxl.load_workbook(filename=file_path)
        else:
            self.wb = openpyxl.Workbook()

        if APARTMENTS_WORKSHEET not in self.wb:
            self.wb.active.title = APARTMENTS_WORKSHEET

        self.ws = self.wb[APARTMENTS_WORKSHEET]
        self.ws.cell(row=self.ws.max_row, column=1)  # Make last row active

    def send_notification(self, url, description, area, actual_data):
        super(FileNotifier, self).send_notification(url, description, area, actual_data)
        self.ws.append([datetime.now(), url, description, area, actual_data])
        self.f.write("{}, {}, {}, {}\n".format(url, description, area, actual_data))

    def finalize(self):
        super(FileNotifier, self).finalize()
        self.wb.save(self.file_path)
